from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, session, send_file
from werkzeug.utils import secure_filename

from app import db
from sqlalchemy import func

from app.models import (
    VerifierUser,
    RecitationInput,
    QuranSurah,
    QuranAyah,
    ErrorDetails,
    RecitationWordDetails,
    Reciter,
    SurahAudio
)
from app.quran_processor import process_recitation, load_surahs_from_db
from flask_mail import Message
from . import mail

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display

import os
import subprocess
import uuid
import threading
import io
import datetime
import time
from datetime import datetime as dt, date, timedelta
from decimal import Decimal
main = Blueprint("main", __name__)


# ============================================================
# Cache Surahs
# ============================================================
_surahs_cache = {}
_surahs_lock = threading.Lock()


def get_surahs_cache():
    global _surahs_cache
    if not _surahs_cache:
        with _surahs_lock:
            if not _surahs_cache:
                _surahs_cache = load_surahs_from_db(db.session)
    return _surahs_cache


# ============================================================
# Media helpers
# ============================================================
def detect_input_type(filename):
    ext = os.path.splitext(filename or "")[1].lower()

    audio_exts = [".mp3", ".wav", ".m4a", ".aac", ".ogg", ".flac"]
    video_exts = [".mp4", ".mov", ".avi", ".mkv", ".webm"]

    if ext in video_exts:
        return "video"
    if ext in audio_exts:
        return "audio"
    return "file"


def get_media_duration(path):
    """
    Returns media duration in seconds using ffprobe.
    Works for audio and video files if ffmpeg/ffprobe is installed.
    """
    try:
        result = subprocess.run(
            [
                "ffprobe",
                "-v", "error",
                "-show_entries", "format=duration",
                "-of", "default=noprint_wrappers=1:nokey=1",
                path
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=20
        )

        duration = result.stdout.strip()
        if duration:
            return int(float(duration))
    except Exception:
        pass

    return None


def format_duration(seconds):
    if seconds is None:
        return "—"

    try:
        seconds = int(float(seconds))
    except Exception:
        return "—"

    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    secs = seconds % 60

    if hours:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"

    return f"{minutes:02d}:{secs:02d}"


def get_saved_original_name(path_or_link):
    """
    Safe fallback display name for history/results.
    If filepathorlink is a local path, show only the filename, never the Windows path.
    """
    if not path_or_link:
        return "—"

    text_value = str(path_or_link)

    if text_value.startswith("http://") or text_value.startswith("https://"):
        return text_value

    return os.path.basename(text_value)


def clean_display_filename(filename):
    """
    Keep the user's original visible name clean.
    This is for display only, not for saving on disk.
    """
    if not filename:
        return "—"

    filename = str(filename).strip()
    filename = os.path.basename(filename.replace("\\", "/"))

    return filename or "—"


def get_storage_filename(path_or_link):
    """
    Return the UUID/stored filename from the saved file path.
    Example: C:/.../uploads/files/uuid.mp3 -> uuid.mp3
    """
    if not path_or_link:
        return ""

    text_value = str(path_or_link)

    if text_value.startswith("http://") or text_value.startswith("https://"):
        return text_value

    return os.path.basename(text_value)


def set_model_attr_if_exists(obj, attr, value):
    """
    Set model attribute only if the current SQLAlchemy model has this column/attribute.
    This prevents crashes if the DB/model does not include original_filename yet.
    """
    if hasattr(obj, attr):
        setattr(obj, attr, value)


# ============================================================
# Static Pages
# ============================================================
@main.route('/')
def home():
    users_count = VerifierUser.query.count()
    recitations_count = RecitationInput.query.count()

    # المستخدمين
    if users_count > 100:
        users_count_display = "+99"
    else:
        users_count_display = str(users_count)

    # التلاوات
    if recitations_count > 500:
        recitations_count_display = "+500"
    else:
        recitations_count_display = str(recitations_count)

    return render_template(
        'landing.html',
        users_count_display=users_count_display,
        recitations_count_display=recitations_count_display
    )


@main.route("/contact", methods=["GET", "POST"])
def contact():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        subject = request.form.get("subject", "").strip()
        message_text = request.form.get("message", "").strip()

        try:
            receiver = os.getenv("CONTACT_RECEIVER") or os.getenv("MAIL_DEFAULT_SENDER")
            msg = Message(
                subject=f"[Ta'yaqan Contact] {subject}",
                recipients=[receiver],
                body=(
                    f"Name: {name}\n"
                    f"Email: {email}\n"
                    f"Subject: {subject}\n\n"
                    f"Message:\n{message_text}\n"
                ),
                reply_to=email if email else None
            )
            mail.send(msg)
            flash("تم إرسال رسالتك بنجاح ✅", "success")
        except Exception as e:
            flash(f"تعذر إرسال الرسالة ❌ — {str(e)}", "error")

        return redirect(url_for("main.contact"))

    return render_template("contact.html")


@main.route("/about")
def about():
    return render_template("about.html")


@main.route("/listen")
def listen():
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    wanted_reciters = [
        "عبد الباسط عبد الصمد",
        "سعد الغامدي",
        "ماهر المعيقلي",
        "ياسر الدوسري"
    ]

    reciters = (
        Reciter.query
        .join(SurahAudio, SurahAudio.reciterid == Reciter.reciterid)
        .filter(Reciter.recitername.in_(wanted_reciters))
        .group_by(Reciter.reciterid, Reciter.recitername)
        .order_by(
            db.case(
                (Reciter.recitername == "عبد الباسط عبد الصمد", 1),
                else_=2
            ),
            Reciter.reciterid.asc()
        )
        .all()
    )

    return render_template("listen.html", reciters=reciters)


@main.route("/listen/<int:reciter_id>")
def listen_reciter(reciter_id):
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    reciter = Reciter.query.get_or_404(reciter_id)

    surah_audios = (
        db.session.query(SurahAudio, QuranSurah)
        .join(QuranSurah, QuranSurah.surahid == SurahAudio.surahid)
        .filter(SurahAudio.reciterid == reciter_id)
        .order_by(QuranSurah.surahid.asc())
        .all()
    )

    return render_template(
        "listen_surahs.html",
        reciter=reciter,
        surah_audios=surah_audios
    )
@main.route("/upload", methods=["GET"])
def upload():
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    modal_message = request.args.get("modal_message", "")
    modal_type = request.args.get("modal_type", "")

    return render_template(
        "upload.html",
        modal_message=modal_message,
        modal_type=modal_type
    )


# ============================================================
# Upload File
# ============================================================
@main.route("/upload/file", methods=["POST"])
def file_verify():
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    f = request.files.get("recitation_file")
    if not f or f.filename.strip() == "":
        flash("رجاءً اختار ملف أولًا ❌", "error")
        return redirect(url_for("main.upload"))

    uploads_dir = os.path.join(current_app.root_path, "static", "uploads", "files")
    os.makedirs(uploads_dir, exist_ok=True)

    file_id = str(uuid.uuid4())

    # الاسم الأصلي للعرض فقط، لا نستخدمه للحفظ حتى لا يصير تعارض أو مشاكل مسارات
    display_original_name = clean_display_filename(f.filename)

    # اسم آمن لاستخراج الامتداد فقط
    safe_original_name = secure_filename(f.filename)
    ext = os.path.splitext(safe_original_name)[1].lower()

    stored_filename = f"{file_id}{ext}"
    saved_path = os.path.join(uploads_dir, stored_filename)

    f.save(saved_path)

    input_type = detect_input_type(display_original_name)

    # If the user uploads a video, extract audio for processing but keep the source type as video.
    if input_type == "video":
        audio_path = os.path.join(uploads_dir, f"{file_id}.mp3")

        try:
            subprocess.run(
                [
                    "ffmpeg",
                    "-y",
                    "-i", saved_path,
                    "-vn",
                    "-acodec", "libmp3lame",
                    "-q:a", "2",
                    audio_path
                ],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            saved_path = audio_path
            stored_filename = f"{file_id}.mp3"
        except subprocess.CalledProcessError:
            flash("تعذر استخراج الصوت من الفيديو ❌", "error")
            return redirect(url_for("main.upload"))

    time.sleep(0.3)

    if not os.path.exists(saved_path) or os.path.getsize(saved_path) == 0:
        flash("فشل رفع الملف ❌ حاول مجددًا.", "error")
        return redirect(url_for("main.upload"))

    session["_pending_path"] = saved_path
    session["_pending_type"] = input_type
    session["_pending_original_name"] = display_original_name
    session["_pending_stored_filename"] = stored_filename
    session["_pending_duration"] = get_media_duration(saved_path)

    return redirect(url_for("main.process_pending"))


# ============================================================
# YouTube Upload
# ============================================================
@main.route("/upload/youtube", methods=["POST"])
def youtube_verify():
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    youtube_url = request.form.get("youtube_url", "").strip()
    if not youtube_url:
        flash("الرجاء إدخال رابط يوتيوب", "error")
        return redirect(url_for("main.upload"))

    downloads_dir = os.path.join(current_app.root_path, "static", "uploads", "youtube")
    os.makedirs(downloads_dir, exist_ok=True)

    file_id = str(uuid.uuid4())
    output_template = os.path.join(downloads_dir, f"{file_id}.%(ext)s")

    import sys
    import shutil

    ytdlp = shutil.which("yt-dlp") or shutil.which("yt-dlp.exe")

    if not ytdlp:
        scripts_dir = os.path.join(os.path.dirname(sys.executable), "Scripts")
        for name in ["yt-dlp.exe", "yt-dlp"]:
            candidate = os.path.join(scripts_dir, name)
            if os.path.exists(candidate):
                ytdlp = candidate
                break

    if not ytdlp:
        ytdlp = os.path.join(
            os.environ.get("APPDATA", ""),
            "Python",
            f"Python{sys.version_info.major}{sys.version_info.minor}",
            "Scripts",
            "yt-dlp.exe"
        )

    if not os.path.exists(ytdlp):
        flash("yt-dlp غير مثبت ❌ شغّلي: pip install yt-dlp", "error")
        return redirect(url_for("main.upload"))

    commands = [
        [
            ytdlp,
            "-x",
            "--audio-format", "mp3",
            "--audio-quality", "0",
            "--cookies-from-browser", "chrome",
            "--no-playlist",
            "--socket-timeout", "20",
            "-o", output_template,
            youtube_url
        ],
        [
            ytdlp,
            "-x",
            "--audio-format", "mp3",
            "--audio-quality", "0",
            "--no-playlist",
            "--socket-timeout", "20",
            "-o", output_template,
            youtube_url
        ]
    ]

    downloaded = False

    for cmd in commands:
        try:
            print("⏳ Downloading YouTube audio...")
            subprocess.run(
                cmd,
                check=True,
                timeout=120
            )
            downloaded = True
            break

        except subprocess.TimeoutExpired:
            print("❌ yt-dlp timeout")
            flash(
                "استغرق تحميل الرابط وقتًا طويلًا ❌ جرّبي رابطًا آخر أو ملفًا صوتيًا مباشرًا.",
                "error"
            )
            return redirect(url_for("main.upload"))

        except subprocess.CalledProcessError as e:
            print(f"❌ yt-dlp failed: {e}")
            continue

    if not downloaded:
        return redirect(url_for(
            "main.upload",
            modal_type="error",
            modal_message="تعذر تحميل رابط اليوتيوب. قد يكون الفيديو غير متاح أو يحتوي على قيود من يوتيوب. يرجى تجربة رابط آخر أو رفع الملف مباشرة."
    ))

    mp3_path = None
    for fname in os.listdir(downloads_dir):
        if fname.startswith(file_id):
            mp3_path = os.path.join(downloads_dir, fname)
            break

    if not mp3_path or not os.path.exists(mp3_path) or os.path.getsize(mp3_path) == 0:
        flash("لم يتم العثور على ملف صوتي صالح بعد التحميل ❌", "error")
        return redirect(url_for("main.upload"))

    print("✅ YouTube downloaded:", mp3_path)

    youtube_title = "رابط يوتيوب"
    try:
        title_result = subprocess.run(
            [ytdlp, "--get-title", "--no-playlist", youtube_url],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=30
        )
        if title_result.stdout.strip():
            youtube_title = title_result.stdout.strip().splitlines()[0]
    except Exception:
        youtube_title = youtube_url

    session["_pending_path"] = mp3_path
    session["_pending_type"] = "youtube"
    session["_pending_original_name"] = clean_display_filename(youtube_title)
    session["_pending_stored_filename"] = get_storage_filename(mp3_path)
    session["_pending_duration"] = get_media_duration(mp3_path)

    return redirect(url_for("main.process_pending"))


# ============================================================
# Process
# ============================================================
@main.route("/process", methods=["POST"])
def process():
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    source_path = request.form.get("source_path", "").strip()
    source_type = request.form.get("source_type", "file")

    if not source_path:
        flash("لم يُحدَّد مصدر صوت ❌", "error")
        return redirect(url_for("main.upload"))

    rec = RecitationInput(
        verifierid=session["user_id"],
        inputtype=source_type,
        filepathorlink=source_path,
        verificationstatus=False,
        processingdate=None,
    )

    set_model_attr_if_exists(rec, "original_filename", get_saved_original_name(source_path))
    set_model_attr_if_exists(rec, "source_name", get_saved_original_name(source_path))
    set_model_attr_if_exists(rec, "stored_filename", get_storage_filename(source_path))
    set_model_attr_if_exists(rec, "duration_sec", get_media_duration(source_path))

    db.session.add(rec)
    db.session.commit()

    input_id = rec.inputid

    def _run(app, iid, src):
        with app.app_context():
            _process_and_save(iid, src)

    threading.Thread(
        target=_run,
        args=(current_app._get_current_object(), input_id, source_path),
        daemon=True
    ).start()

    return redirect(url_for("main.progress", input_id=input_id))


# ============================================================
# Process Pending
# ============================================================
@main.route("/process_pending", methods=["GET"])
def process_pending():
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    source_path = session.pop("_pending_path", None)
    source_type = session.pop("_pending_type", "file")
    source_name = session.pop("_pending_original_name", None)
    stored_filename = session.pop("_pending_stored_filename", None)
    duration_sec = session.pop("_pending_duration", None)

    if not source_path:
        flash("لم يُحدَّد مصدر صوت ❌", "error")
        return redirect(url_for("main.upload"))

    rec = RecitationInput(
        verifierid=session["user_id"],
        inputtype=source_type,
        filepathorlink=source_path,
        verificationstatus=False,
        processingdate=None,
    )

    set_model_attr_if_exists(rec, "duration_sec", duration_sec)
    set_model_attr_if_exists(rec, "original_filename", clean_display_filename(source_name) if source_name else get_saved_original_name(source_path))
    set_model_attr_if_exists(rec, "source_name", clean_display_filename(source_name) if source_name else get_saved_original_name(source_path))
    set_model_attr_if_exists(rec, "stored_filename", stored_filename or get_storage_filename(source_path))
    db.session.add(rec)
    db.session.commit()

    input_id = rec.inputid

    def _run(app, iid, src):
        with app.app_context():
            _process_and_save(iid, src)

    threading.Thread(
        target=_run,
        args=(current_app._get_current_object(), input_id, source_path),
        daemon=True
    ).start()

    return redirect(url_for("main.progress", input_id=input_id))


# ============================================================
# Progress
# ============================================================
@main.route("/progress/<int:input_id>")
def progress(input_id):
    if not session.get("user_id"):
        return redirect(url_for("auth.login"))

    RecitationInput.query.filter_by(
        inputid=input_id,
        verifierid=session["user_id"]
    ).first_or_404()

    return render_template("progress.html", input_id=input_id)


@main.route("/progress/status/<int:input_id>")
def progress_status(input_id):
    if not session.get("user_id"):
        return {
            "done": False,
            "error": True,
            "error_message": "غير مسجل دخول"
        }, 401

    rec = RecitationInput.query.filter_by(
        inputid=input_id,
        verifierid=session["user_id"]
    ).first_or_404()

    response = {
        "done": False,
        "error": False,
        "percent": rec.progress_percent or 0,
        "step": rec.progress_step or "",
        "label": rec.progress_label or ""
    }

    if rec.verificationstatus is True:
        response["done"] = True
        response["redirect_url"] = url_for("main.results", input_id=input_id)
        return response

    if bool(getattr(rec, "audioissue", False)):
        response["done"] = True
        response["audio_issue"] = True
        response["audio_issue_message"] = (
            getattr(rec, "audioissuemessage", "")
            or "تعذر تحليل التسجيل بسبب مشكلة في الصوت."
        )
        response["redirect_url"] = url_for("main.upload")
        return response

    if rec.verificationstatus is False and rec.progress_step == "error":
        response["done"] = True
        response["error"] = True
        response["error_message"] = (
            rec.progress_label
            or "حصل خطأ أثناء المعالجة ❌"
        )
        return response

    return response


# ============================================================
# Processing Helpers
# ============================================================
def update_progress(rec, percent, step, label):
    rec.progress_percent = percent
    rec.progress_step = step
    rec.progress_label = label
    db.session.commit()


def set_processing_time(rec, started_at):
    """
    Save the real processing duration in seconds.
    It starts when _process_and_save begins and ends when the result, audio issue,
    or error state is saved.
    """
    if not started_at:
        return

    if hasattr(rec, "processing_seconds"):
        try:
            rec.processing_seconds = max(0, int(round(time.time() - started_at)))
        except Exception:
            rec.processing_seconds = None


def set_audio_issue(rec, reason, message):
    rec.processingdate = dt.utcnow()
    rec.verificationstatus = False
    rec.audioissue = True
    rec.audioissuereason = reason
    rec.audioissuemessage = message
    update_progress(rec, 100, "audio_issue", message)


def extract_word_text(word_obj):
    return (
        word_obj.get("word")
        or word_obj.get("spoken")
        or word_obj.get("original")
        or word_obj.get("expected")
        or ""
    ).strip()


def normalize_token(token):
    if not token:
        return ""

    replacements = {
        "أ": "ا",
        "إ": "ا",
        "آ": "ا",
        "ى": "ي",
        "ة": "ه",
        "ؤ": "و",
        "ئ": "ي",
        "ٱ": "ا",
    }

    token = str(token).strip()
    for old, new in replacements.items():
        token = token.replace(old, new)

    token = token.replace("ـ", "")
    return token


def remove_consecutive_repetitions(words, max_phrase_len=10):
    """
    يحذف التكرار المتتابع من التقرير فقط.
    مثال: قل قل هو الله أحد -> يحذف قل الثانية.
    """
    cleaned = []
    i = 0
    tokens = [normalize_token(extract_word_text(w)) for w in words]

    while i < len(words):
        repeated = False
        max_len = min(max_phrase_len, (len(words) - i) // 2)

        for phrase_len in range(max_len, 0, -1):
            first = tokens[i:i + phrase_len]
            second = tokens[i + phrase_len:i + (2 * phrase_len)]

            if first and first == second and all(first):
                cleaned.extend(words[i:i + phrase_len])
                i += 2 * phrase_len
                repeated = True
                break

        if not repeated:
            cleaned.append(words[i])
            i += 1

    return cleaned


def detect_group_recitation(words, min_phrase_len=4, max_phrase_len=20, min_repeats=2):
    """
    يكشف القراءات الجماعية أو نمط قارئ يقرأ ثم أطفال/جماعة يكررون بعده.
    إذا وجد مقطع طويل مكرر بشكل متتابع، نعرض ويندو بأن هذا النوع غير مدعوم.
    """
    tokens = [
        normalize_token(extract_word_text(w))
        for w in words
        if extract_word_text(w)
    ]

    if len(tokens) < min_phrase_len * min_repeats:
        return False

    for phrase_len in range(max_phrase_len, min_phrase_len - 1, -1):
        if len(tokens) < phrase_len * min_repeats:
            continue

        for i in range(0, len(tokens) - (phrase_len * min_repeats) + 1):
            first = tokens[i:i + phrase_len]
            second = tokens[i + phrase_len:i + (2 * phrase_len)]

            if first == second and all(first):
                return True

    return False


def detect_bad_echo_result(result, surah_data):
    words = result.get("words", [])
    if not words:
        return True, "no_words_detected"

    expected_words_count = 0
    for ayah in surah_data.get("ayahs", []):
        text = ayah.get("text", "") if isinstance(ayah, dict) else str(ayah)
        expected_words_count += len(text.split())

    if expected_words_count == 0:
        return False, ""

    total_words = len(words)
    missing_count = sum(1 for w in words if w.get("missing"))
    extra_count = sum(1 for w in words if w.get("extra"))
    corrected_count = sum(1 for w in words if w.get("corrected"))

    coverage_ratio = total_words / expected_words_count
    missing_ratio = missing_count / expected_words_count
    corrected_ratio = corrected_count / max(total_words, 1)
    extra_ratio = extra_count / max(total_words, 1)

    last_ayah = surah_data.get("ayahcount", 1)
    reached_ayahs = [
        w.get("ayah", 1)
        for w in words
        if not w.get("missing") and not w.get("extra")
    ]

    max_reached_ayah = max(reached_ayahs) if reached_ayahs else 1
    tail_not_reached = max_reached_ayah < max(1, int(last_ayah * 0.75))

    bad_score = 0
    reasons = []

    if coverage_ratio < 0.70:
        bad_score += 2
        reasons.append("low_coverage")

    if tail_not_reached:
        bad_score += 2
        reasons.append("tail_not_reached")

    if missing_ratio > 0.25:
        bad_score += 2
        reasons.append("high_missing_ratio")

    if corrected_ratio > 0.25:
        bad_score += 1
        reasons.append("high_correction_ratio")

    if extra_ratio > 0.20:
        bad_score += 1
        reasons.append("high_extra_ratio")

    if bad_score >= 3:
        return True, ", ".join(reasons)

    return False, ""


# ============================================================
# Internal Processing Function
# ============================================================
def _process_and_save(input_id: int, source_path: str):
    processing_started_at = time.time()

    rec = RecitationInput.query.get(input_id)
    if not rec:
        return

    audio_issue_message = (
        "يبدو أن التسجيل يحتوي على صدى أو تلاوة جماعية، وهذا النوع غير مدعوم حاليًا، لذلك لا يمكن تحليل التلاوة."
    )

    group_recitation_message = (
        "يبدو أن التسجيل يحتوي على صدى أو تلاوة جماعية، وهذا النوع غير مدعوم حاليًا، لذلك لا يمكن تحليل التلاوة."
    )

    try:
        update_progress(rec, 5, "prepare", "جاري تجهيز الملف...")

        db_surahs = get_surahs_cache()
        update_progress(rec, 15, "load_data", "جاري تحميل بيانات القرآن...")

        result = process_recitation(
            source_path,
            db_surahs,
            progress_callback=lambda percent, step, label: update_progress(rec, percent, step, label)
        )

        if result.get("stop_early"):
            set_processing_time(rec, processing_started_at)
            set_audio_issue(
                rec=rec,
                reason=result.get("audio_issue_reason", "audio_issue"),
                message=audio_issue_message
            )
            return

        surah_number = result["surah"]["number"]
        surah_data = db_surahs[surah_number]
        original_words = result.get("words", [])

        if detect_group_recitation(original_words):
            set_processing_time(rec, processing_started_at)
            set_audio_issue(
                rec=rec,
                reason="group_recitation",
                message=group_recitation_message
            )
            return

        words = remove_consecutive_repetitions(original_words)
        result["words"] = words

        is_bad_audio, bad_reason = detect_bad_echo_result(result, surah_data)
        if is_bad_audio:
            set_processing_time(rec, processing_started_at)
            set_audio_issue(
                rec=rec,
                reason="echo_or_bad_audio",
                message=audio_issue_message
            )
            print(f"⚠️ Audio issue detected for input_id={input_id}: {bad_reason}")
            return

        ayah_ids = surah_data.get("ayah_ids", [])

        rec.surahid = surah_number
        rec.processingdate = dt.utcnow()
        rec.totalwords = len(words)
        rec.startayah = 1
        rec.endayah = surah_data["ayahcount"]
        rec.audioissue = False
        rec.audioissuereason = ""
        rec.audioissuemessage = ""

        RecitationWordDetails.query.filter_by(inputid=input_id).delete()

        update_progress(rec, 96, "save_words", "جاري حفظ تفاصيل الكلمات...")

        correct_counter = 0

        for idx, w in enumerate(words):
            ayah_num = w.get("ayah", 1)
            ayah_id = ayah_ids[ayah_num - 1] if 0 < ayah_num <= len(ayah_ids) else None

            if w.get("missing"):
                status = "ناقص"
            elif w.get("extra"):
                status = "زائد"
            elif w.get("source") == "db_verse":
                status = "صحيح"
            elif w.get("corrected") and w.get("source") == "db_corrected":
                orig = w.get("original", "")
                expected = w.get("word", "")
                conf = w.get("confidence", 0)

                if orig and expected:
                    from rapidfuzz import fuzz
                    similarity = fuzz.ratio(orig, expected)

                    if similarity >= 80 or conf < 0.6 or w.get("asr_noise"):
                        status = "صحيح"
                    else:
                        status = "تحريف"
                else:
                    status = "تحريف"
            else:
                status = "صحيح"

            if status == "صحيح":
                correct_counter += 1

            if status == "ناقص":
                expected_word = w["word"]
                spoken_word = None
            elif status == "زائد":
                expected_word = None
                spoken_word = w["word"]
            elif status == "تحريف":
                expected_word = w["word"]
                spoken_word = w.get("original", "")
            else:
                expected_word = w["word"]
                spoken_word = w.get("original", w["word"]) if w.get("corrected") else w["word"]

            db.session.add(RecitationWordDetails(
                inputid=input_id,
                referenceayahid=ayah_id,
                ayahnumber=ayah_num,
                word_index=idx,
                expected_word=expected_word,
                spoken_word=spoken_word,
                status=status,
                starttime=Decimal(str(w.get("start", 0))),
                endtime=Decimal(str(w.get("end", 0))),
                notes=w.get("source", ""),
            ))

        rec.correctwords = correct_counter

        update_progress(rec, 99, "finalizing", "جاري إعداد النتيجة النهائية...")

        set_processing_time(rec, processing_started_at)
        rec.verificationstatus = True
        db.session.commit()

        update_progress(rec, 100, "done", "اكتملت المعالجة ✅")

    except Exception as e:
        db.session.rollback()

        rec = RecitationInput.query.get(input_id)
        if not rec:
            return

        rec.verificationstatus = False
        rec.processingdate = dt.utcnow()
        set_processing_time(rec, processing_started_at)

        try:
            rec.progress_percent = 100
            rec.progress_step = "error"
            rec.progress_label = "حدث خطأ أثناء المعالجة ❌"
            db.session.commit()
        except Exception:
            db.session.rollback()

        import traceback
        print(f"❌ خطأ input_id={input_id}: {e}")
        traceback.print_exc()
        current_app.logger.error(f"خطأ input_id={input_id}: {e}")


# ============================================================
# History
# ============================================================
@main.route("/history")
def history():
    user_id = session.get("user_id")
    if not user_id:
        return redirect(url_for("auth.login"))

    rows = (
        db.session.query(
            RecitationInput,
            QuranSurah.surahname.label("surahname"),
            func.count(RecitationWordDetails.wordid)
            .filter(RecitationWordDetails.status != "صحيح")
            .label("errors_count"),
        )
        .outerjoin(QuranSurah, QuranSurah.surahid == RecitationInput.surahid)
        .outerjoin(RecitationWordDetails, RecitationWordDetails.inputid == RecitationInput.inputid)
        .filter(RecitationInput.verifierid == user_id)
        .filter(RecitationInput.audioissue != True)
        .group_by(RecitationInput.inputid, QuranSurah.surahname)
        .order_by(RecitationInput.processingdate.desc().nullslast())
        .all()
    )

    input_ids = [r.RecitationInput.inputid for r in rows]
    errors_map = {}

    if input_ids:
        word_errs = (
            db.session.query(
                RecitationWordDetails.inputid,
                RecitationWordDetails.ayahnumber,
                RecitationWordDetails.status,
                RecitationWordDetails.expected_word,
                RecitationWordDetails.spoken_word,
            )
            .filter(RecitationWordDetails.inputid.in_(input_ids))
            .filter(RecitationWordDetails.status.in_(["ناقص", "زائد", "تحريف"]))
            .order_by(
                RecitationWordDetails.inputid.asc(),
                RecitationWordDetails.ayahnumber.asc().nullslast(),
                RecitationWordDetails.word_index.asc().nullslast(),
            )
            .all()
        )

        for inputid, ayahnumber, status, expected_word, spoken_word in word_errs:
            if status == "ناقص":
                msg = f"نقص كلمة: {expected_word or ''}".strip()
            elif status == "زائد":
                msg = f"زيادة كلمة: {spoken_word or ''}".strip()
            else:
                msg = (
                    f"تحريف: المتوقع '{expected_word or ''}' "
                    f"— المنطوق '{spoken_word or ''}'"
                ).strip()

            errors_map.setdefault(inputid, []).append({
                "ayahnumber": ayahnumber,
                "errortype": status,
                "mismatchedtext": msg,
            })

    display_name_map = {}
    duration_map = {}
    processing_duration_map = {}

    for row in rows:
        rec_obj = row.RecitationInput
        input_id = rec_obj.inputid

        display_name = (
            getattr(rec_obj, "original_filename", None)
            or getattr(rec_obj, "source_name", None)
            or get_saved_original_name(rec_obj.filepathorlink)
        )
        display_name = clean_display_filename(display_name)

        display_name_map[input_id] = display_name
        duration_map[input_id] = format_duration(getattr(rec_obj, "duration_sec", None))
        processing_duration_map[input_id] = format_duration(getattr(rec_obj, "processing_seconds", None))

    return render_template(
        "history.html",
        rows=rows,
        errors_map=errors_map,
        display_name_map=display_name_map,
        duration_map=duration_map,
        processing_duration_map=processing_duration_map
    )


# ============================================================
# Results
# ============================================================
@main.route("/results/<int:input_id>")
def results(input_id):
    user_id = session.get("user_id")
    is_admin = session.get("is_admin", False)
    admin_id = session.get("admin_id")

    if not user_id and not admin_id:
        return redirect(url_for("auth.login"))

    if is_admin:
        rec = RecitationInput.query.filter_by(inputid=input_id).first_or_404()
    else:
        rec = RecitationInput.query.filter_by(
            inputid=input_id,
            verifierid=user_id
        ).first_or_404()

    errors = (
        db.session.query(ErrorDetails, QuranAyah.ayahnumber)
        .join(QuranAyah, QuranAyah.ayahid == ErrorDetails.referenceayahid)
        .filter(ErrorDetails.inputid == rec.inputid)
        .all()
    )

    word_details = (
        RecitationWordDetails.query
        .filter_by(inputid=rec.inputid)
        .order_by(
            RecitationWordDetails.ayahnumber.asc().nullslast(),
            RecitationWordDetails.word_index.asc().nullslast(),
            RecitationWordDetails.starttime.asc().nullslast(),
            RecitationWordDetails.wordid.asc()
        )
        .all()
    )

    correct_count = sum(1 for w in word_details if w.status == "صحيح")
    missing_count = sum(1 for w in word_details if w.status == "ناقص")
    extra_count = sum(1 for w in word_details if w.status == "زائد")
    wrong_count = sum(1 for w in word_details if w.status == "تحريف")
    total_count = len(word_details)
    errors_count = missing_count + extra_count + wrong_count
    is_ok = errors_count == 0

    audio_url = None

    if rec and rec.filepathorlink:
        abs_path = rec.filepathorlink

        try:
            static_root = os.path.join(current_app.root_path, "static")
            abs_norm = os.path.normpath(abs_path)
            static_norm = os.path.normpath(static_root)

            if abs_norm.startswith(static_norm) and os.path.exists(abs_norm):
                rel_path = os.path.relpath(abs_norm, static_root).replace("\\", "/")
                audio_url = url_for("static", filename=rel_path)
        except Exception:
            audio_url = None

    # ===== Back button target =====
    # If admin opens results from admin pages, keep the admin inside the admin flow.
    admin_view = request.args.get("admin_view") == "1"
    admin_user_id = request.args.get("user_id", type=int)
    admin_surah_id = request.args.get("surah_id", type=int)

    back_url = url_for("main.history")
    back_label = "العودة للوحة التحكم"

    if is_admin:
        if admin_view and admin_user_id and admin_surah_id:
            back_url = url_for(
                "main.admin_user_surah",
                user_id=admin_user_id,
                surah_id=admin_surah_id
            )
            back_label = "العودة لتحققات السورة"
        elif admin_view and admin_user_id:
            back_url = url_for("main.admin_user_detail", user_id=admin_user_id)
            back_label = "العودة لتفاصيل المستخدم"
        else:
            # Fallback for old admin links that do not send admin_view/user_id/surah_id.
            referrer = request.referrer or ""
            if f"/admin/users/{rec.verifierid}/surah/{rec.surahid}" in referrer:
                back_url = url_for(
                    "main.admin_user_surah",
                    user_id=rec.verifierid,
                    surah_id=rec.surahid
                )
                back_label = "العودة لتحققات السورة"
            elif f"/admin/users/{rec.verifierid}" in referrer:
                back_url = url_for("main.admin_user_detail", user_id=rec.verifierid)
                back_label = "العودة لتفاصيل المستخدم"
            elif "/admin/verifications" in referrer:
                back_url = url_for("main.admin_verifications")
                back_label = "العودة للتلاوات والتحقق"
            elif "/admin/reports" in referrer:
                back_url = url_for("main.admin_reports")
                back_label = "العودة للتقارير العامة"
            else:
                back_url = url_for("main.admin_dashboard")
                back_label = "العودة للوحة الإدارة"

    return render_template(
        "results.html",
        rec=rec,
        errors=errors,
        word_details=word_details,
        total_count=total_count,
        correct_count=correct_count,
        missing_count=missing_count,
        extra_count=extra_count,
        wrong_count=wrong_count,
        errors_count=errors_count,
        is_ok=is_ok,
        audio_url=audio_url,
        source_display_name=clean_display_filename(getattr(rec, "original_filename", None) or getattr(rec, "source_name", None) or get_saved_original_name(rec.filepathorlink)),
        duration_display=format_duration(getattr(rec, "duration_sec", None)),
        processing_duration_display=format_duration(getattr(rec, "processing_seconds", None)),
        audio_issue=bool(getattr(rec, "audioissue", False)),
        audio_issue_message=getattr(rec, "audioissuemessage", ""),
        audio_issue_reason=getattr(rec, "audioissuereason", ""),
        back_url=back_url,
        back_label=back_label
    )


@main.route("/results_stub/<int:input_id>")
def results_stub(input_id):
    return redirect(url_for("main.results", input_id=input_id))


# ============================================================
# Download PDF
# ============================================================
@main.route("/results/<int:input_id>/download-pdf")
def download_pdf(input_id):
    user_id = session.get("user_id")
    is_admin = session.get("is_admin", False)
    admin_id = session.get("admin_id")

    if not user_id and not admin_id:
        return redirect(url_for("auth.login"))

    if is_admin:
        rec = RecitationInput.query.filter_by(inputid=input_id).first_or_404()
    else:
        rec = RecitationInput.query.filter_by(
            inputid=input_id,
            verifierid=user_id
        ).first_or_404()

    word_details = (
        RecitationWordDetails.query
        .filter_by(inputid=rec.inputid)
        .order_by(
            RecitationWordDetails.ayahnumber.asc().nullslast(),
            RecitationWordDetails.word_index.asc().nullslast(),
        )
        .all()
    )

    correct_count = sum(1 for w in word_details if w.status == "صحيح")
    missing_count = sum(1 for w in word_details if w.status == "ناقص")
    extra_count = sum(1 for w in word_details if w.status == "زائد")
    wrong_count = sum(1 for w in word_details if w.status == "تحريف")
    errors_count = missing_count + extra_count + wrong_count
    is_ok = errors_count == 0

    font_path = os.path.join(current_app.root_path, "static", "fonts", "Amiri-Regular.ttf")
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", font_path))

    def ar(text):
        return get_display(arabic_reshaper.reshape(str(text)))

    purple = colors.HexColor("#64449a")
    green = colors.HexColor("#22c55e")
    red = colors.HexColor("#ef4444")
    orange = colors.HexColor("#f59e0b")
    gray_bg = colors.HexColor("#f9fafb")
    border = colors.HexColor("#e5e7eb")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    def s(size=11, color=colors.black, align="RIGHT"):
        return ParagraphStyle(
            name=f"s{size}{id(color)}{align}",
            fontName="Amiri",
            fontSize=size,
            textColor=color,
            alignment={"RIGHT": 2, "CENTER": 1, "LEFT": 0}[align],
            leading=size * 1.6,
        )

    story = []

    story.append(Paragraph(ar("تقرير نتيجة التحقق"), s(20, purple, "CENTER")))
    story.append(Spacer(1, 0.3 * cm))

    surah_name = rec.surah.surahname if hasattr(rec, "surah") and rec.surah else str(rec.surahid)
    story.append(Paragraph(ar(f"سورة {surah_name}"), s(13, colors.HexColor("#6b7280"), "CENTER")))
    story.append(Spacer(1, 0.2 * cm))

    now_str = datetime.datetime.now().strftime("%Y-%m-%d  %H:%M")
    story.append(Paragraph(ar(f"التاريخ: {now_str}"), s(10, colors.HexColor("#9ca3af"), "CENTER")))
    story.append(Spacer(1, 0.6 * cm))

    result_color = green if is_ok else red
    result_bg = colors.HexColor("#f0fdf4") if is_ok else colors.HexColor("#fff1f2")
    result_text = ar("سليمة") if is_ok else ar("غير سليمة")
    sub_text = ar("التلاوة صحيحة دون أي أخطاء") if is_ok else ar(f"تم رصد {errors_count} خطأ في التلاوة")

    rt = Table(
        [
            [Paragraph(result_text, s(22, result_color, "CENTER"))],
            [Paragraph(sub_text, s(11, result_color, "CENTER"))],
        ],
        colWidths=[17 * cm],
    )
    rt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), result_bg),
        ("BOX", (0, 0), (-1, -1), 1.5, result_color),
        ("TOPPADDING", (0, 0), (-1, -1), 14),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
    ]))
    story.append(rt)
    story.append(Spacer(1, 0.6 * cm))

    stats = [
        (ar("صحيح"), correct_count, colors.HexColor("#eefbf2"), green),
        (ar("خطأ"), wrong_count, colors.HexColor("#fdeff0"), red),
        (ar("زائد"), extra_count, colors.HexColor("#fff6e8"), orange),
        (ar("ناقص"), missing_count, colors.HexColor("#fff6e8"), orange),
    ]

    stats_row = [[
        Table(
            [
                [Paragraph(label, s(12, col, "CENTER"))],
                [Paragraph(str(val), s(24, col, "CENTER"))],
            ],
            colWidths=[3.8 * cm],
            style=TableStyle([
                ("BACKGROUND", (0, 0), (-1, -1), bg),
                ("BOX", (0, 0), (-1, -1), 1, col),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ])
        )
        for label, val, bg, col in stats
    ]]

    st = Table(stats_row, colWidths=[4.1 * cm] * 4, hAlign="CENTER")
    st.setStyle(TableStyle([("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    story.append(st)
    story.append(Spacer(1, 0.8 * cm))

    story.append(Paragraph(ar("تفاصيل الكلمات"), s(14, purple)))
    story.append(Spacer(1, 0.3 * cm))

    status_colors = {
        "صحيح": (colors.HexColor("#ecfdf3"), colors.HexColor("#166534")),
        "تحريف": (colors.HexColor("#fff1f2"), red),
        "زائد": (colors.HexColor("#fff6e8"), orange),
        "ناقص": (colors.HexColor("#fefce8"), colors.HexColor("#854d0e")),
    }

    header = [
        Paragraph(ar("ملاحظات"), s(10, colors.white, "CENTER")),
        Paragraph(ar("الوقت"), s(10, colors.white, "CENTER")),
        Paragraph(ar("الحالة"), s(10, colors.white, "CENTER")),
        Paragraph(ar("الكلمة"), s(10, colors.white, "CENTER")),
        Paragraph(ar("آية"), s(10, colors.white, "CENTER")),
    ]
    table_data = [header]

    def fmt_time(sec):
        if sec is None:
            return "--"
        t = int(float(sec))
        return f"{t // 3600:02d}:{(t % 3600) // 60:02d}:{t % 60:02d}"

    for w in word_details:
        if w.status == "ناقص":
            word = w.expected_word or "---"
        elif w.status == "زائد":
            word = w.spoken_word or "---"
        else:
            word = w.expected_word or w.spoken_word or "---"

        note = w.notes if w.notes and str(w.notes).strip() else "لا يوجد"

        row = [
            Paragraph(ar(note), s(9, colors.HexColor("#6b7280"), "CENTER")),
            Paragraph(ar(fmt_time(w.starttime)), s(9, colors.black, "CENTER")),
            Paragraph(ar(w.status), s(9, colors.black, "CENTER")),
            Paragraph(ar(word), s(10, colors.black, "CENTER")),
            Paragraph(str(w.ayahnumber or "-"), s(10, colors.black, "CENTER")),
        ]
        table_data.append(row)

    wt = Table(
        table_data,
        colWidths=[4.5 * cm, 3 * cm, 3 * cm, 4 * cm, 2.5 * cm],
        repeatRows=1
    )
    ts = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), purple),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, gray_bg]),
        ("BOX", (0, 0), (-1, -1), 0.5, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, border),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ])

    for i, w in enumerate(word_details, start=1):
        bg, fc = status_colors.get(w.status, (colors.white, colors.black))
        ts.add("BACKGROUND", (0, i), (-1, i), bg)
        ts.add("TEXTCOLOR", (2, i), (2, i), fc)

    wt.setStyle(ts)
    story.append(wt)

    story.append(Spacer(1, 1 * cm))
    story.append(Paragraph(
        ar("Ta'yaqan - منصة التحقق من التلاوة"),
        s(9, colors.HexColor("#9ca3af"), "CENTER")
    ))

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"report_surah_{rec.surahid}.pdf",
        mimetype="application/pdf"
    )


# ============================================================
# User Reports
# ============================================================
@main.route("/reports")
def reports():
    user_id = session.get("user_id")
    if not user_id:
        return redirect(url_for("auth.login"))

    period = request.args.get("period", "week")
    now = dt.utcnow()
    since_map = {"week": 7, "month": 30, "3months": 90, "year": 365}
    since = now - timedelta(days=since_map[period]) if period in since_map else None

    q = (
        RecitationInput.query
        .filter_by(verifierid=user_id)
        .filter(RecitationInput.audioissue != True)
    )

    if since:
        q = q.filter(RecitationInput.processingdate >= since)

    all_recs = q.all()
    input_ids = [r.inputid for r in all_recs]
    rec_map = {r.inputid: r for r in all_recs}

    total_files = len(all_recs)

    all_words = (
        RecitationWordDetails.query
        .filter(RecitationWordDetails.inputid.in_(input_ids))
        .all()
        if input_ids else []
    )

    errors_per_input = {}
    for w in all_words:
        if w.status != "صحيح":
            errors_per_input[w.inputid] = errors_per_input.get(w.inputid, 0) + 1

    files_with_errors = sum(1 for v in errors_per_input.values() if v > 0)
    files_without_errors = total_files - files_with_errors
    total_errors = sum(errors_per_input.values())

    missing_count = sum(1 for w in all_words if w.status == "ناقص")
    extra_count = sum(1 for w in all_words if w.status == "زائد")
    wrong_count = sum(1 for w in all_words if w.status == "تحريف")

    surah_counter = {}
    for r in all_recs:
        if r.surahid:
            surah_counter[r.surahid] = surah_counter.get(r.surahid, 0) + 1

    top_surah_id = max(surah_counter, key=surah_counter.get) if surah_counter else None
    top_surah = QuranSurah.query.get(top_surah_id) if top_surah_id else None
    top_surah_name = top_surah.surahname if top_surah else "—"

    error_type_counts = {
        "ناقص": missing_count,
        "زائد": extra_count,
        "تحريف": wrong_count
    }

    most_common_error = (
        max(error_type_counts, key=error_type_counts.get)
        if any(error_type_counts.values())
        else None
    )

    most_common_error_ar = {
        "ناقص": "نقص كلمات",
        "زائد": "زيادة كلمات",
        "تحريف": "تحريف كلمات"
    }.get(most_common_error, "—")

    most_common_error_count = error_type_counts.get(most_common_error, 0)

    word_errors = []
    for w in all_words:
        if w.status != "صحيح":
            rec_obj = rec_map.get(w.inputid)
            surah_obj = QuranSurah.query.get(rec_obj.surahid) if rec_obj and rec_obj.surahid else None
            word = w.expected_word if w.status == "ناقص" else w.spoken_word

            if word:
                word_errors.append((
                    word,
                    w.status,
                    w.inputid,
                    surah_obj.surahname if surah_obj else "—"
                ))

    word_agg = {}
    for word, status, inputid, surah_name in word_errors:
        key = (word, status, surah_name)
        if key not in word_agg:
            word_agg[key] = {"count": 0, "files": set()}

        word_agg[key]["count"] += 1
        word_agg[key]["files"].add(inputid)

    top_words = sorted(
        [
            {
                "word": k[0],
                "status": k[1],
                "surah": k[2],
                "count": v["count"],
                "files": len(v["files"])
            }
            for k, v in word_agg.items()
        ],
        key=lambda x: x["count"],
        reverse=True
    )[:10]

    audio_count = sum(
        1 for r in all_recs
        if r.inputtype
        and r.inputtype.lower() in ["file", "audio"]
    )

    youtube_count = sum(
        1 for r in all_recs
        if r.inputtype
        and "youtube" in r.inputtype.lower()
    )

    video_count = sum(
        1 for r in all_recs
        if r.inputtype
        and "video" in r.inputtype.lower()
        and "youtube" not in r.inputtype.lower()
    )

    other_file_count = sum(
        1 for r in all_recs
        if r.inputtype
        and r.inputtype.lower() not in ["file", "audio"]
        and "youtube" not in r.inputtype.lower()
        and "video" not in r.inputtype.lower()
    )
    audio_count += other_file_count

    source_real_total = audio_count + youtube_count + video_count

    source_distribution = [
        {
            "type": "ملفات صوتية",
            "count": audio_count,
            "pct": round(audio_count / source_real_total * 100) if source_real_total else 0
        },
        {
            "type": "روابط يوتيوب",
            "count": youtube_count,
            "pct": round(youtube_count / source_real_total * 100) if source_real_total else 0
        },
        {
            "type": "ملفات فيديو",
            "count": video_count,
            "pct": round(video_count / source_real_total * 100) if source_real_total else 0
        },
    ]

    top_source = (
        max(source_distribution, key=lambda x: x["count"])
        if source_real_total
        else {"type": "لا توجد مصادر بعد", "count": 0, "pct": 0}
    )

    surah_agg = {}
    for r in all_recs:
        if r.surahid:
            if r.surahid not in surah_agg:
                surah_agg[r.surahid] = {
                    "count": 0,
                    "errors": 0,
                    "files_with_errors": set()
                }

            surah_agg[r.surahid]["count"] += 1

    for w in all_words:
        if w.status != "صحيح":
            rec_obj = rec_map.get(w.inputid)
            if rec_obj and rec_obj.surahid and rec_obj.surahid in surah_agg:
                surah_agg[rec_obj.surahid]["errors"] += 1
                surah_agg[rec_obj.surahid]["files_with_errors"].add(w.inputid)

    top_surahs = []
    for surah_id, data in sorted(
        surah_agg.items(),
        key=lambda x: x[1]["count"],
        reverse=True
    )[:5]:
        surah_obj = QuranSurah.query.get(surah_id)

        if surah_obj:
            top_surahs.append({
                "name": surah_obj.surahname,
                "verifications": data["count"],
                "errors": data["errors"],
                "files_with_errors": len(data["files_with_errors"]),
            })

    recent_recs_q = (
        RecitationInput.query
        .filter_by(verifierid=user_id)
        .filter(RecitationInput.verificationstatus == True)
        .filter(RecitationInput.audioissue != True)
        .order_by(RecitationInput.processingdate.desc())
        .limit(4)
        .all()
    )

    recent_verifications = []
    for r in recent_recs_q:
        surah_obj = QuranSurah.query.get(r.surahid) if r.surahid else None
        err_count = errors_per_input.get(r.inputid)

        if err_count is None:
            err_count = (
                RecitationWordDetails.query
                .filter_by(inputid=r.inputid)
                .filter(RecitationWordDetails.status != "صحيح")
                .count()
            )

        recent_verifications.append({
            "inputid": r.inputid,
            "surah_name": surah_obj.surahname if surah_obj else "—",
            "start_ayah": r.startayah or 1,
            "end_ayah": r.endayah or "—",
            "date": r.processingdate.strftime("%Y-%m-%d") if r.processingdate else "—",
            "time": r.processingdate.strftime("%H:%M") if r.processingdate else "—",
            "source_type": r.inputtype or "file",
            "errors": err_count,
        })

    return render_template(
        "reports.html",
        period=period,
        total_files=total_files,
        files_with_errors=files_with_errors,
        files_without_errors=files_without_errors,
        total_errors=total_errors,
        top_surah_name=top_surah_name,
        missing_count=missing_count,
        extra_count=extra_count,
        wrong_count=wrong_count,
        most_common_error_ar=most_common_error_ar,
        most_common_error_count=most_common_error_count,
        top_words=top_words,
        source_distribution=source_distribution,
        source_real_total=source_real_total,
        top_source=top_source,
        top_surahs=top_surahs,
        recent_verifications=recent_verifications,
    )


@main.route("/reports/pdf")
def reports_pdf():
    period = request.args.get("period", "week")
    return redirect(url_for("main.reports", period=period))


# ============================================================
@main.route("/reports/excel")
def reports_excel():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    user_id = session.get("user_id")
    if not user_id:
        return redirect(url_for("auth.login"))

    period = request.args.get("period", "week")
    now = dt.utcnow()
    since_map = {"week": 7, "month": 30, "3months": 90, "year": 365}
    since = now - timedelta(days=since_map[period]) if period in since_map else None

    q = (
        RecitationInput.query
        .filter_by(verifierid=user_id)
        .filter(RecitationInput.audioissue != True)
    )

    if since:
        q = q.filter(RecitationInput.processingdate >= since)

    all_recs = q.order_by(RecitationInput.processingdate.desc()).all()
    input_ids = [r.inputid for r in all_recs]

    all_words = (
        RecitationWordDetails.query
        .filter(RecitationWordDetails.inputid.in_(input_ids))
        .all()
        if input_ids else []
    )

    errors_per_input = {}
    for w in all_words:
        if w.status != "صحيح":
            errors_per_input[w.inputid] = errors_per_input.get(w.inputid, 0) + 1

    total_files = len(all_recs)
    files_with_errors = sum(1 for v in errors_per_input.values() if v > 0)
    files_without_errors = total_files - files_with_errors
    total_errors = sum(errors_per_input.values())
    missing_count = sum(1 for w in all_words if w.status == "ناقص")
    extra_count = sum(1 for w in all_words if w.status == "زائد")
    wrong_count = sum(1 for w in all_words if w.status == "تحريف")

    wb = Workbook()

    ws = wb.active
    ws.title = "الإحصائيات العامة"
    ws.sheet_view.rightToLeft = True

    purple = "8E83A9"
    white = "FFFFFF"
    light = "F7F2E9"
    green = "D1FAE5"
    red = "FEE2E2"

    ws.merge_cells("A1:B1")
    ws["A1"] = "تقرير التحقق الشامل"
    ws["A1"].font = Font(bold=True, color=purple, name="Arial", size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:B2")
    ws["A2"] = f"الفترة: {period} | تاريخ التصدير: {dt.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(color="888888", name="Arial", size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["البيان", "القيمة"])

    for cell in ws[4]:
        cell.fill = PatternFill("solid", start_color=purple)
        cell.font = Font(bold=True, color=white, name="Arial")
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ["إجمالي الملفات المتحقق منها", total_files],
        ["ملفات سليمة بدون أخطاء", files_without_errors],
        ["ملفات بها أخطاء", files_with_errors],
        ["مجموع الأخطاء الكلي", total_errors],
        ["أخطاء نقص كلمات", missing_count],
        ["أخطاء تحريف كلمات", wrong_count],
        ["أخطاء زيادة كلمات", extra_count],
    ]

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=light)
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    ws2 = wb.create_sheet("آخر التحققات")
    ws2.sheet_view.rightToLeft = True
    headers = ["السورة", "من آية", "إلى آية", "التاريخ", "الوقت", "المصدر", "عدد الأخطاء", "الحالة"]
    ws2.append(headers)

    for cell in ws2[1]:
        cell.fill = PatternFill("solid", start_color=purple)
        cell.font = Font(bold=True, color=white, name="Arial")
        cell.alignment = Alignment(horizontal="center")

    for r in all_recs[:50]:
        surah_obj = QuranSurah.query.get(r.surahid) if r.surahid else None
        err_count = errors_per_input.get(r.inputid, 0)
        source_type = (r.inputtype or "file").lower()

        if "youtube" in source_type:
            source_label = "يوتيوب"
        elif "video" in source_type:
            source_label = "فيديو"
        else:
            source_label = "صوتي"

        ws2.append([
            surah_obj.surahname if surah_obj else "—",
            r.startayah or 1,
            r.endayah or "—",
            r.processingdate.strftime("%Y-%m-%d") if r.processingdate else "—",
            r.processingdate.strftime("%H:%M") if r.processingdate else "—",
            source_label,
            err_count,
            "سليمة" if err_count == 0 else "بها أخطاء"
        ])

        for cell in ws2[ws2.max_row]:
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=green if err_count == 0 else red)

    for i, width in enumerate([20, 10, 10, 15, 10, 12, 15, 15], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"taqrir_{period}_{dt.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



@main.route("/reports/pdf-export")
def reports_pdf_export():
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import arabic_reshaper
    from bidi.algorithm import get_display

    user_id = session.get("user_id")
    if not user_id:
        return redirect(url_for("auth.login"))

    # ===== نفس حسابات الصفحة =====
    period = request.args.get("period", "month")
    now = dt.utcnow()
    since_map = {"week": 7, "month": 30, "3months": 90, "year": 365}
    since = now - timedelta(days=since_map[period]) if period in since_map else None

    q = RecitationInput.query.filter_by(verifierid=user_id)
    if since:
        q = q.filter(RecitationInput.processingdate >= since)
    all_recs = q.all()

    input_ids = [r.inputid for r in all_recs]
    total_files = len(all_recs)

    all_words = RecitationWordDetails.query.filter(
        RecitationWordDetails.inputid.in_(input_ids)
    ).all() if input_ids else []

    errors_per_input = {}
    for w in all_words:
        if w.status != "صحيح":
            errors_per_input[w.inputid] = errors_per_input.get(w.inputid, 0) + 1

    files_with_errors = sum(1 for v in errors_per_input.values() if v > 0)
    files_without_errors = total_files - files_with_errors
    total_errors = sum(errors_per_input.values())

    missing_count = sum(1 for w in all_words if w.status == "ناقص")
    extra_count = sum(1 for w in all_words if w.status == "زائد")
    wrong_count = sum(1 for w in all_words if w.status == "تحريف")

    # أكثر سورة
    surah_counter = {}
    for r in all_recs:
        if r.surahid:
            surah_counter[r.surahid] = surah_counter.get(r.surahid, 0) + 1

    top_surah_id = max(surah_counter, key=surah_counter.get) if surah_counter else None
    top_surah = QuranSurah.query.get(top_surah_id) if top_surah_id else None
    top_surah_name = top_surah.surahname if top_surah else "—"

    # ===== الخط العربي =====
    font_path = os.path.join(current_app.root_path, "static", "fonts", "Amiri-Regular.ttf")
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", font_path))

    def ar(text):
        return get_display(arabic_reshaper.reshape(str(text)))

    # ===== ألوان مثل الموقع =====
    PURPLE = colors.HexColor("#8E83A9")
    GREEN = colors.HexColor("#10B981")
    RED = colors.HexColor("#EF4444")
    GRAY = colors.HexColor("#6B7280")

    def style(size=12, color=colors.black, align="RIGHT"):
     return ParagraphStyle(
        name=f"s{size}",
        fontName="Amiri",
        fontSize=size,
        textColor=color,
        alignment={"RIGHT": 2, "CENTER": 1, "LEFT": 0}[align],
        leading=size * 1.6,
        wordWrap="RTL",
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                            rightMargin=2*cm, leftMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    story = []

    # ===== العنوان =====
    story.append(Paragraph(ar("تقارير التحقق الشاملة"), style(18, PURPLE, "CENTER")))
    story.append(Spacer(1, 0.2*cm))
    story.append(Paragraph(ar("إحصائيات وتحليل شامل لجميع عمليات التحقق"), style(10, GRAY, "CENTER")))
    story.append(Spacer(1, 0.6*cm))

    # ===== الكروت (أهم تحسين 🔥) =====
    def card(title, value, color):
        value_text = ar(value) if isinstance(value, str) else str(value)

        return Table(
         [
            [Paragraph(value_text, style(18, color, "CENTER"))],
            [Paragraph(ar(title), style(10, GRAY, "CENTER"))],
        ],
        colWidths=[4*cm],
        style=TableStyle([
            ("BOX", (0,0), (-1,-1), 1, color),
            ("TOPPADDING", (0,0), (-1,-1), 10),
            ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ])
    )

    cards = Table([[
        card("إجمالي الملفات", total_files, PURPLE),
        card("ملفات سليمة", files_without_errors, GREEN),
        card("ملفات بها أخطاء", files_with_errors, RED),
        card("الأكثر تحققاً", top_surah_name, PURPLE),
    ]], colWidths=[4*cm]*4)

    story.append(cards)
    story.append(Spacer(1, 0.7*cm))

    # ===== جدول الأخطاء =====
    story.append(Paragraph(ar("توزيع أنواع الأخطاء"), style(14, PURPLE)))
    story.append(Spacer(1, 0.3*cm))

    data = [
        [
            Paragraph(ar("العدد"), style(12, colors.white, "CENTER")),
            Paragraph(ar("النوع"), style(12, colors.white, "CENTER")),
        ],
        [
            Paragraph(str(missing_count), style(12, colors.black, "CENTER")),
            Paragraph(ar("نقص كلمات"), style(12, colors.black, "RIGHT")),
        ],
        [
            Paragraph(str(wrong_count), style(12, colors.black, "CENTER")),
            Paragraph(ar("تحريف كلمات"), style(12, colors.black, "RIGHT")),
        ],
        [
            Paragraph(str(extra_count), style(12, colors.black, "CENTER")),
            Paragraph(ar("زيادة كلمات"), style(12, colors.black, "RIGHT")),
        ],
        [
            Paragraph(str(total_errors), style(12, GRAY, "CENTER")),
            Paragraph(ar("المجموع"), style(12, GRAY, "RIGHT")),
        ],
    ]

    table = Table(data, colWidths=[4*cm, 8*cm], hAlign="CENTER")

    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), PURPLE),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#D8D2E3")),
        ("TOPPADDING", (0,0), (-1,-1), 9),
        ("BOTTOMPADDING", (0,0), (-1,-1), 9),
    ]))

    story.append(table)

    # ===== الفوتر =====
    story.append(Spacer(1, 1*cm))
    story.append(Paragraph(ar("Ta'yaqan - منصة التحقق من التلاوة"), style(9, GRAY, "CENTER")))

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"taqrir_{period}.pdf",
        mimetype="application/pdf"
    )



# Admin
# ============================================================
@main.route("/admin/dashboard")
def admin_dashboard():
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    from sqlalchemy import func as sqlfunc
    from sqlalchemy.orm import joinedload

    total_users = VerifierUser.query.count()
    total_verifications = RecitationInput.query.count()
    total_errors = RecitationWordDetails.query.filter(
        RecitationWordDetails.status != "صحيح"
    ).count()

    today = date.today()
    active_today = db.session.query(
        sqlfunc.count(sqlfunc.distinct(RecitationInput.verifierid))
    ).filter(
        sqlfunc.date(RecitationInput.processingdate) == today
    ).scalar() or 0

    recent_verifications = (
        RecitationInput.query
        .options(
            joinedload(RecitationInput.verifier),
            joinedload(RecitationInput.surah)
        )
        .filter(RecitationInput.processingdate.isnot(None))
        .order_by(RecitationInput.processingdate.desc())
        .limit(5)
        .all()
    )

    all_verifications = (
        RecitationInput.query
        .order_by(RecitationInput.processingdate.desc())
        .limit(50)
        .all()
    )

    users = VerifierUser.query.filter_by(is_admin=False).all()

    files_with_errors = (
        db.session.query(sqlfunc.count(sqlfunc.distinct(RecitationWordDetails.inputid)))
        .filter(RecitationWordDetails.status != "صحيح")
        .scalar() or 0
    )
    files_without_errors = total_verifications - files_with_errors

    common = (
        db.session.query(
            RecitationWordDetails.status,
            sqlfunc.count(RecitationWordDetails.status).label("cnt")
        )
        .filter(RecitationWordDetails.status != "صحيح")
        .group_by(RecitationWordDetails.status)
        .order_by(sqlfunc.count(RecitationWordDetails.status).desc())
        .first()
    )

    error_map = {
        "ناقص": "نقص كلمات",
        "زائد": "زيادة كلمات",
        "تحريف": "تحريف كلمات"
    }
    most_common_error_ar = error_map.get(common[0], common[0]) if common else "—"

    top_surah = (
        db.session.query(
            RecitationInput.surahid,
            sqlfunc.count(RecitationInput.surahid).label("cnt")
        )
        .filter(RecitationInput.surahid.isnot(None))
        .group_by(RecitationInput.surahid)
        .order_by(sqlfunc.count(RecitationInput.surahid).desc())
        .first()
    )

    top_surah_name = "—"
    if top_surah:
        s = QuranSurah.query.get(top_surah[0])
        top_surah_name = s.surahname if s else "—"

    return render_template(
        "admin/admin_dashboard.html",
        total_users=total_users,
        total_verifications=total_verifications,
        total_errors=total_errors,
        active_today=active_today,
        recent_verifications=recent_verifications,
        all_verifications=all_verifications,
        users=users,
        files_with_errors=files_with_errors,
        files_without_errors=files_without_errors,
        most_common_error_ar=most_common_error_ar,
        top_surah_name=top_surah_name,
    )


@main.route("/admin/delete-user/<int:user_id>", methods=["POST"])
def admin_delete_user(user_id):
    if not session.get("is_admin"):
        return {"success": False}, 403

    user = VerifierUser.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()

    return {"success": True}


@main.route("/admin/users")
def admin_users():
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    users = VerifierUser.query.filter_by(is_admin=False).all()
    return render_template("admin/admin_users.html", users=users, now=dt.now())


@main.route("/admin/verifications")
def admin_verifications():
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    verifications = (
        RecitationInput.query
        .order_by(RecitationInput.processingdate.desc())
        .all()
    )
    return render_template("admin/admin_verifications.html", verifications=verifications)


@main.route("/add-test")
def add_test():
    email = "test@tayaqan.com"
    existing = VerifierUser.query.filter_by(verifieremail=email).first()

    if existing:
        return "Already exists ✅"

    db.session.add(VerifierUser(
        verifiername="Test User",
        verifieremail=email,
        verifierpassword="123"
    ))
    db.session.commit()

    return "Inserted ✅"


@main.route("/create-admin")
def create_admin():
    from werkzeug.security import generate_password_hash
    from app.models import Admin

    email = "admin@tayaqan.com"
    password = "Admin12345"

    existing = Admin.query.filter_by(adminemail=email).first()
    if existing:
        return "Admin already exists ✅"

    admin = Admin(
        adminname="Main Admin",
        adminemail=email,
        adminpassword=generate_password_hash(password)
    )

    db.session.add(admin)
    db.session.commit()

    return "Admin created successfully ✅"


@main.route("/admin/reports")
def admin_reports():
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    from datetime import datetime, timedelta
    from sqlalchemy import func as sqlfunc

    period = request.args.get("period", "week")
    now = datetime.utcnow()

    if period == "week":
        since = now - timedelta(days=7)
    elif period == "month":
        since = now - timedelta(days=30)
    elif period == "3months":
        since = now - timedelta(days=90)
    elif period == "year":
        since = now - timedelta(days=365)
    else:
        since = None

    q = RecitationInput.query.filter(RecitationInput.audioissue != True)
    if since:
        q = q.filter(RecitationInput.processingdate >= since)

    all_recs = q.all()
    input_ids = [r.inputid for r in all_recs]

    total_verifications = len(all_recs)
    total_users = VerifierUser.query.count()

    all_words = []
    if input_ids:
        all_words = RecitationWordDetails.query.filter(
            RecitationWordDetails.inputid.in_(input_ids)
        ).all()

    total_errors = sum(1 for w in all_words if w.status != "صحيح")
    missing_count = sum(1 for w in all_words if w.status == "ناقص")
    extra_count = sum(1 for w in all_words if w.status == "زائد")
    wrong_count = sum(1 for w in all_words if w.status == "تحريف")

    file_count = sum(
        1 for r in all_recs
        if r.inputtype
        and "youtube" not in r.inputtype.lower()
        and "video" not in r.inputtype.lower()
    )
    youtube_count = sum(
        1 for r in all_recs
        if r.inputtype and "youtube" in r.inputtype.lower()
    )
    video_count = sum(
        1 for r in all_recs
        if r.inputtype
        and "video" in r.inputtype.lower()
        and "youtube" not in r.inputtype.lower()
    )

    weekly_data = [0] * 7
    for r in all_recs:
        if r.processingdate:
            day = r.processingdate.weekday()
            mapped = {5: 0, 6: 1, 0: 2, 1: 3, 2: 4, 3: 5, 4: 6}
            weekly_data[mapped[day]] += 1

    surah_data = {}
    for r in all_recs:
        if r.surahid:
            if r.surahid not in surah_data:
                surah_data[r.surahid] = {
                    "count": 0,
                    "errors": 0,
                    "files_with_errors": set()
                }
            surah_data[r.surahid]["count"] += 1

    for w in all_words:
        if w.status != "صحيح":
            rec_obj = RecitationInput.query.get(w.inputid)
            if rec_obj and rec_obj.surahid and rec_obj.surahid in surah_data:
                surah_data[rec_obj.surahid]["errors"] += 1
                surah_data[rec_obj.surahid]["files_with_errors"].add(w.inputid)

    top_surahs = []
    for surah_id, data in sorted(
        surah_data.items(),
        key=lambda x: x[1]["count"],
        reverse=True
    )[:6]:
        surah_obj = QuranSurah.query.get(surah_id)
        if surah_obj:
            top_surahs.append({
                "name": surah_obj.surahname,
                "count": data["count"],
                "errors": data["errors"],
                "files_with_errors": len(data["files_with_errors"]),
            })

    all_users = VerifierUser.query.filter_by(is_admin=False).all()
    top_users = sorted(all_users, key=lambda u: len(u.inputs), reverse=True)[:10]

    word_errors = []
    for w in all_words:
        if w.status != "صحيح":
            rec_obj = RecitationInput.query.get(w.inputid)
            surah_obj = QuranSurah.query.get(rec_obj.surahid) if rec_obj and rec_obj.surahid else None
            word = w.expected_word if w.status == "ناقص" else w.spoken_word

            if word:
                word_errors.append((
                    word,
                    w.status,
                    w.inputid,
                    surah_obj.surahname if surah_obj else "—"
                ))

    word_agg = {}
    for word, status, inputid, surah_name in word_errors:
        key = (word, status, surah_name)
        if key not in word_agg:
            word_agg[key] = {"count": 0, "files": set()}

        word_agg[key]["count"] += 1
        word_agg[key]["files"].add(inputid)

    top_words = sorted(
        [
            {
                "word": k[0],
                "status": k[1],
                "surah": k[2],
                "count": v["count"],
                "files": len(v["files"])
            }
            for k, v in word_agg.items()
        ],
        key=lambda x: x["count"],
        reverse=True
    )[:10]

    return render_template(
        "admin/admin_reports.html",
        period=period,
        total_users=total_users,
        total_verifications=total_verifications,
        total_errors=total_errors,
        missing_count=missing_count,
        extra_count=extra_count,
        wrong_count=wrong_count,
        file_count=file_count,
        youtube_count=youtube_count,
        video_count=video_count,
        weekly_data=weekly_data,
        top_surahs=top_surahs,
        top_users=top_users,
        top_words=top_words,
    )


# ============================================================
# Added from current routes.py - admin/listen support
# ============================================================
@main.route("/admin/reports/download/pdf")
def download_admin_report_pdf():
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    period = request.args.get("period", "month")
    now = dt.utcnow()
    since_map = {"week": 7, "month": 30, "3months": 90, "year": 365}
    period_ar = {"week": "أسبوع", "month": "شهر", "3months": "3 أشهر", "year": "عام", "all": "الكل"}
    since = now - timedelta(days=since_map[period]) if period in since_map else None

    q = RecitationInput.query
    if since:
        q = q.filter(RecitationInput.processingdate >= since)

    all_recs = q.all()
    input_ids = [r.inputid for r in all_recs]

    all_words = RecitationWordDetails.query.filter(
        RecitationWordDetails.inputid.in_(input_ids)
    ).all() if input_ids else []

    total_users = VerifierUser.query.filter_by(is_admin=False).count()
    total_verifications = len(all_recs)
    total_errors = sum(1 for w in all_words if w.status != "صحيح")
    missing_count = sum(1 for w in all_words if w.status == "ناقص")
    extra_count = sum(1 for w in all_words if w.status == "زائد")
    wrong_count = sum(1 for w in all_words if w.status == "تحريف")

    surah_counter = {}
    for r in all_recs:
        if r.surahid:
            surah_counter[r.surahid] = surah_counter.get(r.surahid, 0) + 1

    top_surah_name = "—"
    if surah_counter:
        top_id = max(surah_counter, key=surah_counter.get)
        s = QuranSurah.query.get(top_id)
        top_surah_name = s.surahname if s else "—"

    font_path = os.path.join(current_app.root_path, "static", "fonts", "Amiri-Regular.ttf")
    if "Amiri" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont("Amiri", font_path))

    def ar(text):
        return get_display(arabic_reshaper.reshape(str(text)))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.6 * cm,
        leftMargin=1.6 * cm,
        topMargin=1.6 * cm,
        bottomMargin=1.6 * cm
    )

    PURPLE = colors.HexColor("#8E83A9")
    DARK_PURPLE = colors.HexColor("#7A7095")
    SOFT = colors.HexColor("#F7F2E9")
    RED = colors.HexColor("#EF4444")
    RED_SOFT = colors.HexColor("#FEE2E2")
    GREEN_SOFT = colors.HexColor("#D1FAE5")
    BORDER = colors.HexColor("#E7E2DD")

    def style(size=11, color=colors.black, align="CENTER"):
        return ParagraphStyle(
            name=f"s{size}{align}",
            fontName="Amiri",
            fontSize=size,
            textColor=color,
            alignment={"RIGHT": 2, "CENTER": 1, "LEFT": 0}[align],
            leading=size * 1.6
        )

    story = []

    story.append(Paragraph(ar("التقرير الإداري الشامل"), style(22, DARK_PURPLE)))
    story.append(Paragraph(ar(f"الفترة: {period_ar.get(period, period)}  |  تاريخ التصدير: {dt.now().strftime('%Y-%m-%d')}"), style(10, colors.HexColor("#777777"))))
    story.append(Spacer(1, 0.5 * cm))

    cards_data = [[
        Paragraph(ar("إجمالي المستخدمين"), style(10, colors.white)),
        Paragraph(ar("إجمالي التحققات"), style(10, colors.white)),
        Paragraph(ar("إجمالي الأخطاء"), style(10, colors.white)),
        Paragraph(ar("السورة الأكثر تحققًا"), style(10, colors.white)),
    ], [
        Paragraph(str(total_users), style(20)),
        Paragraph(str(total_verifications), style(20)),
        Paragraph(str(total_errors), style(20, RED)),
        Paragraph(ar(top_surah_name), style(16)),
    ]]

    cards = Table(cards_data, colWidths=[4.2 * cm] * 4)
    cards.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
        ("BACKGROUND", (0, 1), (-1, 1), SOFT),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    story.append(cards)
    story.append(Spacer(1, 0.7 * cm))

    story.append(Paragraph(ar("توزيع أنواع الأخطاء"), style(17, DARK_PURPLE)))
    story.append(Spacer(1, 0.25 * cm))

    total_err = missing_count + extra_count + wrong_count or 1

    error_data = [
        [ar("نوع الخطأ"), ar("العدد"), ar("النسبة")],
        [ar("نقص كلمات"), str(missing_count), f"{round(missing_count / total_err * 100)}%"],
        [ar("تحريف كلمات"), str(wrong_count), f"{round(wrong_count / total_err * 100)}%"],
        [ar("زيادة كلمات"), str(extra_count), f"{round(extra_count / total_err * 100)}%"],
    ]

    error_data = [[row[2], row[1], row[0]] for row in error_data]

    err_table = Table(error_data, colWidths=[4 * cm, 4 * cm, 7 * cm])

    # ترتيب الأعمدة بصريًا من اليمين لليسار:
    # نوع الخطأ | العدد | النسبة
    error_data = [[row[2], row[1], row[0]] for row in error_data]

    err_table = Table(error_data, colWidths=[4 * cm, 4 * cm, 7 * cm])
    err_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), PURPLE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), SOFT),
        ("FONTNAME", (0, 0), (-1, -1), "Amiri"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 9),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
    ]))
    story.append(err_table)

    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(ar("Ta'yaqan - منصة التحقق من التلاوة"), style(9, colors.HexColor("#999999"))))

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"admin_report_{period}.pdf",
        mimetype="application/pdf"
    )

@main.route("/admin/reports/download/excel")
def download_admin_report_excel():
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    period = request.args.get("period", "month")
    now = dt.utcnow()
    since_map = {"week": 7, "month": 30, "3months": 90, "year": 365}
    since = now - timedelta(days=since_map[period]) if period in since_map else None

    q = RecitationInput.query
    if since:
        q = q.filter(RecitationInput.processingdate >= since)

    all_recs = q.order_by(RecitationInput.processingdate.desc()).all()
    input_ids = [r.inputid for r in all_recs]

    all_words = RecitationWordDetails.query.filter(
        RecitationWordDetails.inputid.in_(input_ids)
    ).all() if input_ids else []

    total_users = VerifierUser.query.filter_by(is_admin=False).count()
    total_verifications = len(all_recs)
    total_errors = sum(1 for w in all_words if w.status != "صحيح")
    missing_count = sum(1 for w in all_words if w.status == "ناقص")
    extra_count = sum(1 for w in all_words if w.status == "زائد")
    wrong_count = sum(1 for w in all_words if w.status == "تحريف")

    wb = Workbook()
    ws = wb.active
    ws.title = "التقرير الإداري"
    ws.sheet_view.rightToLeft = True

    purple = "8E83A9"
    white = "FFFFFF"
    light = "F7F2E9"

    ws.merge_cells("A1:B1")
    ws["A1"] = "التقرير الإداري الشامل"
    ws["A1"].font = Font(bold=True, color=purple, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["البيان", "القيمة"])

    for cell in ws[3]:
        cell.fill = PatternFill("solid", start_color=purple)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center")

    rows = [
        ["إجمالي المستخدمين", total_users],
        ["إجمالي التحققات", total_verifications],
        ["إجمالي الأخطاء", total_errors],
        ["أخطاء النقص", missing_count],
        ["أخطاء الزيادة", extra_count],
        ["أخطاء التحريف", wrong_count],
    ]

    for row in rows:
        ws.append(row)

    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.fill = PatternFill("solid", start_color=light)
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"admin_report_{period}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
 # ══════════════════════════════════════════════
# أضيفي هذا الـ route في routes.py بعد admin_users مباشرةً
# ══════════════════════════════════════════════

@main.route("/admin/users/<int:user_id>")
def admin_user_detail(user_id):
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    user = VerifierUser.query.get_or_404(user_id)

    # جيب كل تحققات المستخدم
    recs = (
        RecitationInput.query
        .filter_by(verifierid=user_id)
        .filter(RecitationInput.verificationstatus == True)
        .filter(RecitationInput.audioissue != True)
        .order_by(RecitationInput.processingdate.desc())
        .all()
    )

    input_ids = [r.inputid for r in recs]
    total = len(recs)

    # كل الكلمات
    all_words = []
    if input_ids:
        all_words = RecitationWordDetails.query.filter(
            RecitationWordDetails.inputid.in_(input_ids)
        ).all()

    # أخطاء لكل ملف
    errors_per_input = {}
    for w in all_words:
        if w.status != "صحيح":
            errors_per_input[w.inputid] = errors_per_input.get(w.inputid, 0) + 1

    files_with_errors    = sum(1 for v in errors_per_input.values() if v > 0)
    files_without_errors = total - files_with_errors
    total_errors         = sum(errors_per_input.values())
    missing_count        = sum(1 for w in all_words if w.status == "ناقص")
    extra_count          = sum(1 for w in all_words if w.status == "زائد")
    wrong_count          = sum(1 for w in all_words if w.status == "تحريف")

    # آخر نشاط
    last_rec  = recs[0] if recs else None
    last_date = last_rec.processingdate if last_rec else None
    is_active = last_date is not None

    # ══ السور التي تحقق منها المستخدم ══
    surah_agg = {}
    for r in recs:
        if r.surahid:
            if r.surahid not in surah_agg:
                surah_agg[r.surahid] = {
                    "count": 0,
                    "errors": 0,
                    "files_with_errors": set(),
                    "files_ok": 0,
                }
            surah_agg[r.surahid]["count"] += 1
            err = errors_per_input.get(r.inputid, 0)
            surah_agg[r.surahid]["errors"] += err
            if err > 0:
                surah_agg[r.surahid]["files_with_errors"].add(r.inputid)
            else:
                surah_agg[r.surahid]["files_ok"] += 1

    user_surahs = []
    for surah_id, data in sorted(surah_agg.items(), key=lambda x: x[1]["count"], reverse=True):
        surah_obj = QuranSurah.query.get(surah_id)
        if surah_obj:
            user_surahs.append({
                "surah_id"         : surah_id,
                "name"             : surah_obj.surahname,
                "count"            : data["count"],
                "errors"           : data["errors"],
                "files_with_errors": len(data["files_with_errors"]),
                "files_ok"         : data["files_ok"],
            })

    return render_template(
        "admin/admin_user_detail.html",
        user=user,
        total=total,
        files_with_errors=files_with_errors,
        files_without_errors=files_without_errors,
        total_errors=total_errors,
        missing_count=missing_count,
        extra_count=extra_count,
        wrong_count=wrong_count,
        last_date=last_date,
        is_active=is_active,
        user_surahs=user_surahs,
        errors_per_input=errors_per_input,
        recs=recs,
    )


# ══════════════════════════════════════════════
# Route لاستعراض تحققات سورة معينة لمستخدم معين
# أضيفيه بعد admin_user_detail
# ══════════════════════════════════════════════

@main.route("/admin/users/<int:user_id>/surah/<int:surah_id>")
def admin_user_surah(user_id, surah_id):
    if not session.get("is_admin"):
        return redirect(url_for("main.upload"))

    user     = VerifierUser.query.get_or_404(user_id)
    surah    = QuranSurah.query.get_or_404(surah_id)

    recs = (
        RecitationInput.query
        .filter_by(verifierid=user_id, surahid=surah_id)
        .filter(RecitationInput.verificationstatus == True)
        .filter(RecitationInput.audioissue != True)
        .order_by(RecitationInput.processingdate.desc())
        .all()
    )

    input_ids = [r.inputid for r in recs]
    all_words = []
    if input_ids:
        all_words = RecitationWordDetails.query.filter(
            RecitationWordDetails.inputid.in_(input_ids)
        ).all()

    errors_per_input = {}
    for w in all_words:
        if w.status != "صحيح":
            errors_per_input[w.inputid] = errors_per_input.get(w.inputid, 0) + 1

    verifications = []
    for r in recs:
        err = errors_per_input.get(r.inputid, 0)
        src = r.inputtype or "file"
        verifications.append({
            "inputid"    : r.inputid,
            "date"       : r.processingdate.strftime("%Y-%m-%d") if r.processingdate else "—",
            "time"       : r.processingdate.strftime("%H:%M")    if r.processingdate else "—",
            "source_type": src,
            "errors"     : err,
            "status"     : r.verificationstatus,
            "start_ayah" : r.startayah or 1,
            "end_ayah"   : r.endayah   or "—",
        })

    return render_template(
        "admin/admin_user_surah.html",
        user=user,
        surah=surah,
        verifications=verifications,
        errors_per_input=errors_per_input,
    )
